import os
import re
import openpyxl
from typing import Optional, Dict

# Default path; overridden by docker-compose mount
XLSX_PATH = os.environ.get("JOURNAL_XLSX", "/app/data/01fenqu.xlsx")
if not os.path.exists(XLSX_PATH):
    # Fallback to project root for local dev
    fallback = os.path.join(os.path.dirname(__file__), "..", "..", "..", "01fenqu.xlsx")
    if os.path.exists(fallback):
        XLSX_PATH = os.path.abspath(fallback)


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).lower().strip())


def _strip_leading_the(text: str) -> str:
    """Remove leading 'the ' for fuzzy matching (e.g. 'the isme journal' -> 'isme journal')."""
    t = _normalize(text)
    if t.startswith("the "):
        return t[4:]
    return t


def _remove_punctuation(text: str) -> str:
    return re.sub(r"[^\w\s]", "", str(text).lower()).strip()


def _to_float(val) -> Optional[float]:
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str) and val.strip():
        try:
            return float(val)
        except ValueError:
            return None
    return None


class JournalMatcher:
    def __init__(self, xlsx_path: Optional[str] = None):
        self.path = xlsx_path or XLSX_PATH
        self._by_name: Dict[str, Dict] = {}
        self._by_name_nopunct: Dict[str, Dict] = {}
        self._by_issn: Dict[str, Dict] = {}
        self._loaded = False

    def _load(self):
        if self._loaded or not os.path.exists(self.path):
            return
        try:
            wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        except Exception:
            return

        # --- Sheet 1: 2025JCR分区 ---
        jcr_sheet = None
        for name in wb.sheetnames:
            if "JCR" in name:
                jcr_sheet = wb[name]
                break

        if jcr_sheet:
            it = iter(jcr_sheet.iter_rows(values_only=True))
            try:
                header = [str(h).strip() if h else "" for h in next(it)]
            except StopIteration:
                header = []
            idx_name = next((i for i, h in enumerate(header) if "期刊" in h and "名" in h), None)
            idx_issn = next((i for i, h in enumerate(header) if h.upper() == "ISSN"), None)
            # Prefer exact "2024JIF", fallback to any column containing "JIF"
            idx_if = next((i for i, h in enumerate(header) if h == "2024JIF"), None)
            if idx_if is None:
                idx_if = next((i for i, h in enumerate(header) if "JIF" in h.upper()), None)
            idx_q = next((i for i, h in enumerate(header) if h.upper() == "QUARTILE"), None)

            for row in it:
                if not row or not row[0]:
                    continue
                name_val = str(row[idx_name]).strip() if idx_name is not None and row[idx_name] else ""
                issn_val = str(row[idx_issn]).strip() if idx_issn is not None and row[idx_issn] else ""
                if_val = row[idx_if] if idx_if is not None else None
                q_val = str(row[idx_q]).strip().upper() if idx_q is not None and row[idx_q] else ""

                entry = {
                    "journal": name_val,
                    "issn": issn_val,
                    "if": _to_float(if_val),
                    "jcr_quartile": q_val if q_val else None,
                    "xinrui_quartile": None,
                    "is_top": False,
                }

                if name_val:
                    self._by_name[_normalize(name_val)] = entry
                    self._by_name_nopunct[_remove_punctuation(name_val)] = entry
                    # Also index without leading "the" for PubMed matching
                    no_the = _strip_leading_the(name_val)
                    if no_the != _normalize(name_val):
                        self._by_name[no_the] = entry
                        self._by_name_nopunct[_remove_punctuation(no_the)] = entry
                if issn_val:
                    self._by_issn[issn_val] = entry

        # --- Sheet 2: 2025中国科学院分区表 ---
        cas_sheet = None
        for name in wb.sheetnames:
            if "中科院" in name or "科学院" in name or "分区表" in name:
                cas_sheet = wb[name]
                break

        if cas_sheet:
            it = iter(cas_sheet.iter_rows(values_only=True))
            try:
                header = [str(h).strip() if h else "" for h in next(it)]
            except StopIteration:
                header = []
            idx_name = next((i for i, h in enumerate(header) if "期刊" in h and "名称" in h), None)
            idx_xr = next((i for i, h in enumerate(header) if "2025" in h and "分区" in h), None)
            idx_top = next((i for i, h in enumerate(header) if h == "Top"), None)

            for row in it:
                if not row or not row[0]:
                    continue
                name_val = str(row[idx_name]).strip() if idx_name is not None and row[idx_name] else ""
                xr_val = str(row[idx_xr]).strip().upper() if idx_xr is not None and row[idx_xr] else ""
                top_val = str(row[idx_top]).strip() if idx_top is not None and row[idx_top] else ""

                norm = _normalize(name_val)
                nopunct = _remove_punctuation(name_val)

                # Merge into existing entries if found
                target = None
                if norm in self._by_name:
                    target = self._by_name[norm]
                elif nopunct in self._by_name_nopunct:
                    target = self._by_name_nopunct[nopunct]

                if target is None and name_val:
                    target = {
                        "journal": name_val,
                        "issn": "",
                        "if": None,
                        "jcr_quartile": None,
                        "xinrui_quartile": None,
                        "is_top": False,
                    }
                    self._by_name[norm] = target
                    self._by_name_nopunct[nopunct] = target

                if target is not None:
                    if xr_val:
                        target["xinrui_quartile"] = xr_val
                    if top_val == "是":
                        target["is_top"] = True

        wb.close()
        self._loaded = True

    def match(self, journal_name: str, issn: Optional[str] = None) -> Optional[Dict]:
        """Return matched partition info or None."""
        self._load()
        if not journal_name:
            return None

        norm = _normalize(journal_name)
        nopunct = _remove_punctuation(journal_name)

        # 1. Exact match (lowercase, stripped spaces)
        if norm in self._by_name:
            return self._by_name[norm]

        # 2. Punctuation-free match
        if nopunct in self._by_name_nopunct:
            return self._by_name_nopunct[nopunct]

        # 3. ISSN match
        if issn and issn in self._by_issn:
            return self._by_issn[issn]

        # 4. Strip leading "the" and try again (PubMed vs xlsx naming difference)
        norm_no_the = _strip_leading_the(journal_name)
        nopunct_no_the = _remove_punctuation(norm_no_the)
        if norm_no_the in self._by_name:
            return self._by_name[norm_no_the]
        if nopunct_no_the in self._by_name_nopunct:
            return self._by_name_nopunct[nopunct_no_the]

        return None


# Singleton
journal_matcher = JournalMatcher()
